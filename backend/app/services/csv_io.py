from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from app.models import AribaDataset, EXPECTED_COLUMNS, REQUIRED_FILES


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_row(raw_row: dict[str, object], headers: list[str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for header in headers:
        normalized[header] = _normalize_cell(raw_row.get(header, ""))
    return normalized


def read_csv_path(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    encodings = ["utf-8-sig", "latin-1"]
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                headers = [h.strip() for h in (reader.fieldnames or []) if h]
                rows = [_normalize_row(row, headers) for row in reader]
                return headers, rows
        except UnicodeDecodeError as err:
            last_error = err
            continue

    if last_error is not None:
        raise last_error

    return [], []


def read_csv_bytes(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    encodings = ["utf-8-sig", "latin-1"]
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            text = data.decode(encoding)
            handle = StringIO(text)
            reader = csv.DictReader(handle)
            headers = [h.strip() for h in (reader.fieldnames or []) if h]
            rows = [_normalize_row(row, headers) for row in reader]
            return headers, rows
        except UnicodeDecodeError as err:
            last_error = err
            continue

    if last_error is not None:
        raise last_error

    return [], []


def read_xlsx_bytes(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    if first_row is None:
        return [], []

    headers = [_normalize_cell(value) for value in first_row]
    headers = [header for header in headers if header]

    rows: list[dict[str, str]] = []
    for row_values in rows_iter:
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = row_values[index] if index < len(row_values) else ""
            record[header] = _normalize_cell(value)
        rows.append(record)

    return headers, rows


def load_ariba_dataset_from_dir(base_dir: Path) -> tuple[AribaDataset, dict[str, list[str]], list[str]]:
    missing_files: list[str] = []
    headers_by_key: dict[str, list[str]] = {}
    data_by_key: dict[str, list[dict[str, str]]] = {}

    for file_name, key in REQUIRED_FILES.items():
        file_path = base_dir / file_name
        if not file_path.exists():
            missing_files.append(file_name)
            data_by_key[key] = []
            headers_by_key[key] = []
            continue

        headers, rows = read_csv_path(file_path)
        expected = EXPECTED_COLUMNS[key]
        normalized_rows = [_normalize_row(row, expected) for row in rows]

        headers_by_key[key] = headers
        data_by_key[key] = normalized_rows

    dataset = AribaDataset(
        contracts=data_by_key.get("contracts", []),
        contract_documents=data_by_key.get("contract_documents", []),
        contract_content_documents=data_by_key.get("contract_content_documents", []),
        contract_teams=data_by_key.get("contract_teams", []),
        import_projects_parameters=data_by_key.get("import_projects_parameters", []),
    )

    return dataset, headers_by_key, missing_files


def load_ariba_dataset_from_zip_bytes(data: bytes) -> tuple[AribaDataset, dict[str, list[str]], list[str], set[str]]:
    with ZipFile(BytesIO(data)) as zip_file:
        available_names = {name for name in zip_file.namelist() if not name.endswith("/")}
        normalized_names = {name.replace("\\", "/") for name in available_names}

        missing_files: list[str] = []
        headers_by_key: dict[str, list[str]] = {}
        data_by_key: dict[str, list[dict[str, str]]] = {}

        for file_name, key in REQUIRED_FILES.items():
            normalized_file_name = file_name.replace("\\", "/")
            file_variants = {
                normalized_file_name,
                f"./{normalized_file_name}",
            }
            match = next((variant for variant in file_variants if variant in normalized_names), None)
            if match is None:
                suffix = f"/{normalized_file_name}"
                match = next((name for name in normalized_names if name.endswith(suffix)), None)

            if match is None:
                missing_files.append(file_name)
                headers_by_key[key] = []
                data_by_key[key] = []
                continue

            with zip_file.open(match, "r") as handle:
                content = handle.read()
                headers, rows = read_csv_bytes(content)
                expected = EXPECTED_COLUMNS[key]
                normalized_rows = [_normalize_row(row, expected) for row in rows]
                headers_by_key[key] = headers
                data_by_key[key] = normalized_rows

    dataset = AribaDataset(
        contracts=data_by_key.get("contracts", []),
        contract_documents=data_by_key.get("contract_documents", []),
        contract_content_documents=data_by_key.get("contract_content_documents", []),
        contract_teams=data_by_key.get("contract_teams", []),
        import_projects_parameters=data_by_key.get("import_projects_parameters", []),
    )

    return dataset, headers_by_key, missing_files, normalized_names


def load_unified_rows_from_file(filename: str, data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    lower_name = filename.lower()

    if lower_name.endswith(".csv"):
        return read_csv_bytes(data)

    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
        return read_xlsx_bytes(data)

    raise ValueError("Formato nao suportado. Use CSV ou XLSX.")


def dataset_to_csv_string(key: str, rows: list[dict[str, str]]) -> str:
    headers = EXPECTED_COLUMNS[key]
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()

    for row in rows:
        normalized = {header: _normalize_cell(row.get(header, "")) for header in headers}
        writer.writerow(normalized)

    return output.getvalue()
