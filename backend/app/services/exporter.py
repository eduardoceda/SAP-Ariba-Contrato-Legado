from __future__ import annotations

import json
import unicodedata
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from app.models import AribaDataset, ExecutiveSummary, ValidationReport
from app.services.csv_io import dataset_to_csv_string

CLID_TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "data" / "templates" / "clid-item-template.xlsx"

CLID_SHEET_SPECS = (
    {
        "target_title": "Contract Header",
        "aliases": ("Contract Header", "Cabeçalho do contrato"),
        "headers": {
            "Event ID": ("Event ID", "Código do evento"),
            "Title": ("Title", "Título"),
            "Description": ("Description", "Descrição"),
            "Requester": ("Requester", "Solicitante"),
            "Company Name": ("Company Name", "Nome da empresa"),
            "Supplier Name": ("Supplier Name", "Nome do fornecedor"),
            "SupplierID Domain": ("SupplierID Domain", "Domínio do código do fornecedor"),
            "SupplierID Value": ("SupplierID Value", "Valor do código do fornecedor"),
            "Contract Source": ("Contract Source", "Origem do contrato"),
            "Buyer Contract ID": ("Buyer Contract ID", "Código do contrato do Buyer"),
            "Term Type": ("Term Type", "Tipo de condição"),
            "Limit Type": ("Limit Type", "Tipo de limite"),
            "Agreement Date": ("Agreement Date", "Data do contrato"),
            "Effective Date": ("Effective Date", "Data de efetivação"),
            "Expiration Date": ("Expiration Date", "Data de vencimento"),
            "Contract Currency": ("Contract Currency", "Moeda do contrato"),
            "Minimum Amount": ("Minimum Amount", "Valor mínimo"),
            "Maximum Amount": ("Maximum Amount", "Valor máximo"),
            "Reference Document": ("Reference Document", "Documento de referência"),
        },
    },
    {
        "target_title": "Contract Item Information",
        "aliases": ("Contract Item Information", "Info. sobre item do contrato"),
        "headers": {
            "Bundle": ("Bundle", "Pacote"),
            "Item Number": ("Item Number", "Número do item"),
            "Short Name": ("Short Name", "Nome abreviado"),
            "Description": ("Description", "Descrição"),
            "Extended Description": ("Extended Description", "Descrição estendida"),
            "Supplier Part Number": ("Supplier Part Number", "Número de peça do fornecedor"),
            "Unit Of Measure": ("Unit Of Measure", "Unidade de medida"),
            "Unit Price": ("Unit Price", "Preço unitário"),
            "Discount Amount": ("Discount Amount", "Valor do desconto"),
            "Supplier Discount(%)": ("Supplier Discount(%)", "Desconto do fornecedor (%)"),
            "Unit Price Currency": ("Unit Price Currency", "Moeda do preço unitário"),
            "Classification Domain": ("Classification Domain", "Domínio de classificação"),
            "Classification Code": ("Classification Code", "Código de classificação"),
            "Quantity": ("Quantity", "Quantidade"),
            "Minimum Quantity": ("Minimum Quantity", "Quantidade mínima"),
            "Maximum Quantity": ("Maximum Quantity", "Quantidade máxima"),
            "Minimum Amount": ("Minimum Amount", "Valor mínimo"),
            "Maximum Amount": ("Maximum Amount", "Valor máximo"),
            "Manufacturer Name": ("Manufacturer Name", "Nome do fabricante"),
            "Manufacturer Part Number": ("Manufacturer Part Number", "NP fabricante"),
            "Limit Type": ("Limit Type", "Tipo de limite"),
            "Number": ("Number", "Número"),
            "LineType": ("LineType", "LineType"),
            "Item Status": ("Item Status", "Status do item"),
            "External System Line Number": ("External System Line Number", "Número da linha do sistema externo"),
            "Source Event ID": ("Source Event ID", "Código do evento de origem"),
            "Line Item Number": ("Line Item Number", "Número do item de linha"),
        },
    },
    {
        "target_title": "Header Attributes",
        "aliases": ("Header Attributes", "Atributos de cabeçalho"),
        "headers": {
            "Attribute Name": ("Attribute Name", "Nome do atributo"),
            "Attribute Value": ("Attribute Value", "Valor do atributo"),
            "Display Text": ("Display Text", "Texto de exibição"),
            "Type": ("Type", "Tipo"),
            "Description": ("Description", "Descrição"),
            "Table Section Column": ("Table Section Column", "Coluna de seção da tabela"),
        },
    },
    {
        "target_title": "Item Attributes",
        "aliases": ("Item Attributes", "Atributos do item"),
        "headers": {
            "Item Number": ("Item Number", "Número do item"),
            "Attribute Name": ("Attribute Name", "Nome do atributo"),
            "Attribute Value": ("Attribute Value", "Valor do atributo"),
            "Display Text": ("Display Text", "Texto de exibição"),
            "Type": ("Type", "Tipo"),
            "Description": ("Description", "Descrição"),
            "Is Term added from Item Master": (
                "Is Term added from Item Master",
                "A condição é adicionada do mestre do item",
            ),
            "Formula": ("Formula", "Fórmula"),
            "Item Status": ("Item Status", "Status do item"),
        },
    },
)

RESERVED_EXPORT_FILES = {
    "contracts.csv",
    "contractdocuments.csv",
    "contractcontentdocuments.csv",
    "contractteams.csv",
    "importprojectsparameters.csv",
    "validation-report.json",
}

ISSUE_GUIDANCE = {
    "MISSING_FILE": "Inclua o arquivo CSV obrigatório no pacote base.",
    "MISSING_COLUMN": "Ajuste o cabeçalho para conter todas as colunas esperadas.",
    "REQUIRED_FIELD": "Preencha o campo obrigatório na linha indicada.",
    "INVALID_TITLE_SPECIAL_CHARACTERS": "Remova acentos e símbolos do campo Title ou aplique a correção automática.",
    "ARIBA_UNSUPPORTED_MAX_DATE": "Troque 9999-12-31 por 9999-01-01 em EffectiveDate ou ExpirationDate.",
    "INVALID_ID_FORMAT": "Padronize o ContractId conforme a regra definida no cliente.",
    "INVALID_DATE": "Use formato de data YYYY-MM-DD.",
    "INVALID_NUMBER": "Use número decimal válido, sem texto.",
    "INVALID_PARTY_FORMAT": "Use sap: + números, por exemplo sap:0000381965.",
    "MISSING_CONTRACT_REFERENCE": "Garanta que o contrato exista em Contracts.csv.",
    "MISSING_ATTACHMENT": "Confirme se o arquivo está no ZIP e no caminho correto.",
    "UNEXPECTED_FILE_PATH": "Ajuste o caminho para a pasta esperada.",
    "INVALID_ATTACHMENT_FOLDER": "Mova o arquivo para a pasta padrão correspondente.",
    "INVALID_ATTACHMENT_EXTENSION": "Use extensões adequadas para anexo e CLID.",
    "DUPLICATE_CONTRACT_ID": "Remova duplicidade de ContractId.",
    "DUPLICATE_DOCUMENT": "Remova linhas duplicadas de documentos.",
    "DUPLICATE_CONTENT_DOCUMENT": "Remova linhas CLID duplicadas.",
    "DUPLICATE_TEAM_MEMBER": "Mantenha uma linha por contrato, grupo e membro.",
    "IMPORT_PARAMS_ROW_COUNT": "Mantenha exatamente 1 linha em ImportProjectsParameters.",
    "UNEXPECTED_IMPORT_PARAM": "Confirme o valor com o template oficial do cliente.",
    "MISSING_TEAM": "Inclua ao menos 1 membro em ContractTeams.csv para o contrato.",
    "MISSING_REQUIRED_GROUP": "Adicione o grupo obrigatório configurado.",
    "EMPTY_CONTRACTS": "Inclua ao menos um contrato antes de gerar o pacote.",
    "UNREFERENCED_ATTACHMENT": "Arquivo extra no ZIP. Pode ser removido ou referenciado.",
    "INVALID_RULE_REGEX": "Corrija a expressão regular na configuração avançada.",
}


def _issue_guidance(code: str) -> str:
    return ISSUE_GUIDANCE.get(code, "Revise a linha e o campo indicados conforme o template.")


def _normalize_attachment_archive_path(path: str) -> str:
    normalized = path.replace("\\", "/").strip().lstrip("./")
    lowered = normalized.lower()

    for expected_root in ("documentos contratos/", "documentos clid/"):
        root_index = lowered.find(expected_root)
        if root_index >= 0:
            return normalized[root_index:]

    return normalized


def _normalize_label(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.lower().split())


def _translate_clid_sheet_rows(source_workbook, spec: dict[str, object]) -> list[list[object]]:
    aliases = {_normalize_label(alias) for alias in spec["aliases"]}
    source_sheet = next((worksheet for worksheet in source_workbook.worksheets if _normalize_label(worksheet.title) in aliases), None)
    if source_sheet is None:
        return []

    rows = list(source_sheet.iter_rows(values_only=True))
    if not rows:
        return []

    source_headers = list(rows[0])
    header_indexes = {
        _normalize_label(header): index
        for index, header in enumerate(source_headers)
        if _normalize_label(header)
    }

    translated_rows: list[list[object]] = []
    for raw_row in rows[1:]:
        if not any(value not in (None, "") for value in raw_row):
            continue

        next_row: list[object] = []
        for target_header, aliases_for_header in spec["headers"].items():
            value = ""
            for alias in aliases_for_header:
                header_index = header_indexes.get(_normalize_label(alias))
                if header_index is None:
                    continue
                if header_index < len(raw_row):
                    value = raw_row[header_index]
                break
            next_row.append(value)
        translated_rows.append(next_row)

    return translated_rows


def _normalize_clid_workbook_bytes(data: bytes) -> bytes:
    if not CLID_TEMPLATE_PATH.exists():
        return data

    try:
        source_workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
        return data

    translated_by_sheet = {
        spec["target_title"]: _translate_clid_sheet_rows(source_workbook, spec) for spec in CLID_SHEET_SPECS
    }
    if not any(rows for rows in translated_by_sheet.values()):
        return data

    template_workbook = load_workbook(CLID_TEMPLATE_PATH)
    for spec in CLID_SHEET_SPECS:
        target_sheet = template_workbook[spec["target_title"]]
        if target_sheet.max_row > 1:
            target_sheet.delete_rows(2, target_sheet.max_row - 1)
        for row in translated_by_sheet[spec["target_title"]]:
            target_sheet.append(row)

    output = BytesIO()
    template_workbook.save(output)
    return output.getvalue()


def _format_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def build_package_zip(dataset: AribaDataset, report: ValidationReport | None = None, include_report_json: bool = True) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as zip_file:
        zip_file.writestr("Contracts.csv", dataset_to_csv_string("contracts", dataset.contracts))
        zip_file.writestr("ContractDocuments.csv", dataset_to_csv_string("contract_documents", dataset.contract_documents))
        zip_file.writestr(
            "ContractContentDocuments.csv",
            dataset_to_csv_string("contract_content_documents", dataset.contract_content_documents),
        )
        zip_file.writestr("ContractTeams.csv", dataset_to_csv_string("contract_teams", dataset.contract_teams))
        zip_file.writestr(
            "ImportProjectsParameters.csv",
            dataset_to_csv_string("import_projects_parameters", dataset.import_projects_parameters),
        )

        if include_report_json and report is not None:
            zip_file.writestr("validation-report.json", json.dumps(report.model_dump(), indent=2, ensure_ascii=False))

    return buffer.getvalue()


def build_package_zip_with_attachments(
    dataset: AribaDataset,
    attachments_zip_bytes: bytes | None = None,
    report: ValidationReport | None = None,
    include_report_json: bool = True,
) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as zip_file:
        zip_file.writestr("Contracts.csv", dataset_to_csv_string("contracts", dataset.contracts))
        zip_file.writestr("ContractDocuments.csv", dataset_to_csv_string("contract_documents", dataset.contract_documents))
        zip_file.writestr(
            "ContractContentDocuments.csv",
            dataset_to_csv_string("contract_content_documents", dataset.contract_content_documents),
        )
        zip_file.writestr("ContractTeams.csv", dataset_to_csv_string("contract_teams", dataset.contract_teams))
        zip_file.writestr(
            "ImportProjectsParameters.csv",
            dataset_to_csv_string("import_projects_parameters", dataset.import_projects_parameters),
        )

        if attachments_zip_bytes:
            written_attachment_paths: set[str] = set()
            zip_file.writestr("Documentos contratos/", "")
            zip_file.writestr("Documentos CLID/", "")
            with ZipFile(BytesIO(attachments_zip_bytes), "r") as source_zip:
                for entry in source_zip.infolist():
                    name = entry.filename.replace("\\", "/")
                    if name.endswith("/"):
                        continue

                    normalized_name = _normalize_attachment_archive_path(name)

                    base_name = normalized_name.split("/")[-1].lower()
                    if base_name in RESERVED_EXPORT_FILES:
                        continue

                    if normalized_name.lower() in written_attachment_paths:
                        continue

                    entry_bytes = source_zip.read(entry.filename)
                    if normalized_name.lower().startswith("documentos clid/") and normalized_name.lower().endswith(
                        (".xlsx", ".xlsm")
                    ):
                        entry_bytes = _normalize_clid_workbook_bytes(entry_bytes)

                    zip_file.writestr(normalized_name, entry_bytes)
                    written_attachment_paths.add(normalized_name.lower())

        if include_report_json and report is not None:
            zip_file.writestr("validation-report.json", json.dumps(report.model_dump(), indent=2, ensure_ascii=False))

    return buffer.getvalue()


def build_report_xlsx(report: ValidationReport, executive_summary: ExecutiveSummary | None = None) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"

    summary_sheet.append(["Indicador", "Valor"])
    summary_sheet.append(["Apto para carga", "Sim" if report.summary.is_valid else "Nao"])
    summary_sheet.append(["Erros", report.summary.errors])
    summary_sheet.append(["Avisos", report.summary.warnings])
    summary_sheet.append(["Informativos", report.summary.infos])
    summary_sheet.append(["Total de inconsistencias", report.summary.total_issues])
    summary_sheet.append([])
    summary_sheet.append(["Arquivo", "Quantidade de linhas"])

    for key, value in report.record_counts.items():
        summary_sheet.append([key, value])
    _format_sheet(summary_sheet)

    if executive_summary is not None:
        executive_sheet = workbook.create_sheet("Executivo")
        executive_sheet.append(["Indicador", "Valor"])
        executive_sheet.append(["Total de contratos", executive_summary.total_contracts])
        executive_sheet.append(["Contratos prontos para importar", executive_summary.contracts_ready_for_import])
        executive_sheet.append(["Contratos com erros", executive_summary.contracts_with_errors])
        executive_sheet.append(["Contratos com avisos", executive_summary.contracts_with_warnings])
        executive_sheet.append(["Contratos com infos", executive_summary.contracts_with_infos])
        executive_sheet.append(["Anexos de contrato mapeados", executive_summary.mapped_contract_documents])
        executive_sheet.append(["Arquivos CLID mapeados", executive_summary.mapped_clid_documents])
        executive_sheet.append(["Membros de time", executive_summary.team_assignments])
        executive_sheet.append(["Percentual de prontidão (%)", executive_summary.readiness_percent])
        executive_sheet.append(["Recomendação", executive_summary.recommendation])
        _format_sheet(executive_sheet)

    issues_sheet = workbook.create_sheet("Inconsistencias")
    issues_sheet.append(["Severidade", "Codigo", "Mensagem", "Como corrigir", "Arquivo", "Linha", "Campo", "Contrato"])
    for issue in report.issues:
        issues_sheet.append(
            [
                issue.severity,
                issue.code,
                issue.message,
                _issue_guidance(issue.code),
                issue.source_file or "",
                issue.row or "",
                issue.field or "",
                issue.contract_id or "",
            ]
        )
    _format_sheet(issues_sheet)

    contract_sheet = workbook.create_sheet("PorContrato")
    contract_sheet.append(["Contrato", "Erros", "Avisos", "Infos", "Total"])

    by_contract: dict[str, dict[str, int]] = {}
    for issue in report.issues:
        contract_id = issue.contract_id or "(sem contrato)"
        if contract_id not in by_contract:
            by_contract[contract_id] = {"error": 0, "warning": 0, "info": 0, "total": 0}
        by_contract[contract_id][issue.severity] += 1
        by_contract[contract_id]["total"] += 1

    for contract_id in sorted(by_contract.keys()):
        row = by_contract[contract_id]
        contract_sheet.append([contract_id, row["error"], row["warning"], row["info"], row["total"]])
    _format_sheet(contract_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
