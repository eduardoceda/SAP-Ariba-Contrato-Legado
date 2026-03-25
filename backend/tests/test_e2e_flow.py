from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.models import AribaDataset, ExecutiveSummary, ValidationIssue, ValidationReport, ValidationSummary
from app.main import app
from app.services.exporter import build_package_zip_with_attachments, build_report_xlsx
import app.services.run_history as run_history

SAMPLE_DIR = Path(__file__).resolve().parent / "fixtures" / "sample_zip"


def _build_zip_from_paths(base_dir: Path, entries: list[str]) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        for entry in entries:
            source = base_dir / entry
            if source.is_dir():
                for child in source.rglob("*"):
                    if child.is_file():
                        archive.write(child, child.relative_to(base_dir).as_posix())
            elif source.is_file():
                archive.write(source, source.relative_to(base_dir).as_posix())
    return buffer.getvalue()


def _build_full_sample_zip(base_dir: Path) -> bytes:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        for child in base_dir.rglob("*"):
            if child.is_file():
                archive.write(child, child.relative_to(base_dir).as_posix())
    return buffer.getvalue()


def _build_portuguese_clid_workbook() -> bytes:
    workbook = Workbook()

    header_sheet = workbook.active
    header_sheet.title = "Cabeçalho do contrato"
    header_sheet.append(
        [
            "Código do evento",
            "Título",
            "Descrição",
            "Solicitante",
            "Nome da empresa",
            "Nome do fornecedor",
            "Domínio do código do fornecedor",
            "Valor do código do fornecedor",
            "Origem do contrato",
            "Código do contrato do Buyer",
            "Tipo de condição",
            "Tipo de limite",
            "Data do contrato",
            "Data de efetivação",
            "Data de vencimento",
            "Moeda do contrato",
            "Valor mínimo",
            "Valor máximo",
            "Documento de referência",
        ]
    )
    header_sheet.append(
        [
            "Doc123",
            "Contrato teste",
            "Descricao teste",
            "USR001",
            "Empresa X",
            "Fornecedor Y",
            "sap",
            "0001",
            "CW1",
            "LCW4700001278",
            "Item",
            "",
            "2026-03-01",
            "2026-03-01",
            "2026-12-31",
            "BRL",
            "10",
            "100",
            "REF-1",
        ]
    )

    item_sheet = workbook.create_sheet("Info. sobre item do contrato")
    item_sheet.append(
        [
            "Pacote",
            "Número do item",
            "Nome abreviado",
            "Descrição",
            "Descrição estendida",
            "Número de peça do fornecedor",
            "Unidade de medida",
            "Preço unitário",
            "Valor do desconto",
            "Desconto do fornecedor (%)",
            "Moeda do preço unitário",
            "Domínio de classificação",
            "Código de classificação",
            "Quantidade",
            "Quantidade mínima",
            "Quantidade máxima",
            "Valor mínimo",
            "Valor máximo",
            "Nome do fabricante",
            "NP fabricante",
            "Tipo de limite",
            "Número",
            "LineType",
            "Status do item",
            "Número da linha do sistema externo",
            "Número da linha do sistema integrado",
            "Código do evento de origem",
            "Número do item de linha",
        ]
    )
    item_sheet.append(
        [
            "",
            "10",
            "Item curto",
            "Descricao item",
            "Descricao item completa",
            "PART-1",
            "UN",
            "100.00",
            "0",
            "0",
            "BRL",
            "custom",
            "1000000073",
            "1",
            "1",
            "1",
            "10",
            "100",
            "Fabricante X",
            "MFG-1",
            "Item",
            "0001",
            "LineType",
            "Published",
            "EXT-1",
            "INT-1",
            "Doc123",
            "10",
        ]
    )

    header_attr_sheet = workbook.create_sheet("Atributos de cabeçalho")
    header_attr_sheet.append(
        [
            "Nome do atributo",
            "Valor do atributo",
            "Texto de exibição",
            "Tipo",
            "Descrição",
            "Coluna de seção da tabela",
        ]
    )
    header_attr_sheet.append(["HeaderAttr", "Valor", "HeaderAttr", "Texto", "Descricao", ""])

    item_attr_sheet = workbook.create_sheet("Atributos do item")
    item_attr_sheet.append(
        [
            "Número do item",
            "Nome do atributo",
            "Valor do atributo",
            "Texto de exibição",
            "Tipo",
            "Descrição",
            "A condição é adicionada do mestre do item",
            "Fórmula",
            "Status do item",
        ]
    )
    item_attr_sheet.append(["10", "Plant", "1010", "Plant", "MasterData", "Descricao", "No", "", ""])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_e2e_critical_flow(tmp_path, monkeypatch) -> None:
    assert SAMPLE_DIR.exists()
    monkeypatch.setattr(run_history, "HISTORY_FILE", tmp_path / "execution_history.json")

    client = TestClient(app)

    health = client.get("/api/health")
    assert health.status_code == 200
    assert health.json().get("status") == "ok"

    sample_zip = _build_full_sample_zip(SAMPLE_DIR)
    analyze_response = client.post(
        "/api/analyze/upload",
        files={"file": ("sample.zip", sample_zip, "application/zip")},
        data={"profile_name": "Cliente E2E"},
    )
    assert analyze_response.status_code == 200
    analysis_payload = analyze_response.json()
    assert analysis_payload.get("run_id")
    run_id = analysis_payload["run_id"]
    assert analysis_payload["report"]["summary"]["errors"] >= 0

    dataset_json = json.dumps(analysis_payload["dataset"], ensure_ascii=False)
    report_json = json.dumps(analysis_payload["report"], ensure_ascii=False)

    attachments_zip = _build_zip_from_paths(SAMPLE_DIR, ["Documentos contratos", "Documentos CLID"])
    attachments_response = client.post(
        "/api/attachments/validate",
        data={"dataset_json": dataset_json},
        files={"attachments_zip": ("attachments.zip", attachments_zip, "application/zip")},
    )
    assert attachments_response.status_code == 200
    assert "summary" in attachments_response.json()

    export_package_response = client.post(
        "/api/export/package",
        json={
            "dataset": analysis_payload["dataset"],
            "report": analysis_payload["report"],
            "include_report_json": True,
            "run_id": run_id,
        },
    )
    assert export_package_response.status_code == 200
    assert len(export_package_response.content) > 500

    export_with_attachments_response = client.post(
        "/api/export/package-with-attachments",
        data={
            "dataset_json": dataset_json,
            "report_json": report_json,
            "include_report_json": "true",
            "run_id": run_id,
        },
        files={"attachments_zip": ("attachments.zip", attachments_zip, "application/zip")},
    )
    assert export_with_attachments_response.status_code == 200
    assert len(export_with_attachments_response.content) > 500

    export_report_response = client.post(
        "/api/export/report.xlsx",
        json={
            "report": analysis_payload["report"],
            "executive_summary": analysis_payload.get("executive_summary"),
            "run_id": run_id,
        },
    )
    assert export_report_response.status_code == 200
    assert (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        in export_report_response.headers.get("content-type", "")
    )

    runs_response = client.get("/api/runs?limit=10")
    assert runs_response.status_code == 200
    runs = runs_response.json().get("runs", [])
    current_run = next((item for item in runs if item.get("run_id") == run_id), None)
    assert current_run is not None
    assert current_run["profile_name"] == "Cliente E2E"
    assert current_run["artifacts"]["package_zip"] is True
    assert current_run["artifacts"]["package_with_attachments_zip"] is True
    assert current_run["artifacts"]["report_xlsx"] is True


def test_report_xlsx_includes_issue_guidance() -> None:
    report = ValidationReport(
        summary=ValidationSummary(errors=1, warnings=1, infos=0, total_issues=2, is_valid=False),
        issues=[
            ValidationIssue(
                severity="error",
                code="REQUIRED_FIELD",
                message="Campo obrigatório vazio: Owner",
                source_file="Contracts.csv",
                row=2,
                field="Owner",
                contract_id="LCW4700001278",
            ),
            ValidationIssue(
                severity="warning",
                code="UNEXPECTED_FILE_PATH",
                message="Arquivo deveria estar na pasta Documentos contratos/.",
                source_file="ContractDocuments.csv",
                row=3,
                field="File",
                contract_id="LCW4700001278",
            ),
        ],
        record_counts={"contracts": 1, "contract_documents": 1},
    )
    executive_summary = ExecutiveSummary(
        total_contracts=1,
        contracts_ready_for_import=0,
        contracts_with_errors=1,
        contracts_with_warnings=1,
        contracts_with_infos=0,
        mapped_contract_documents=1,
        mapped_clid_documents=0,
        team_assignments=0,
        readiness_percent=0.0,
        recommendation="Ajustar dados obrigatórios antes da carga.",
    )

    workbook = load_workbook(BytesIO(build_report_xlsx(report, executive_summary)))
    assert "Inconsistencias" in workbook.sheetnames

    issues_sheet = workbook["Inconsistencias"]
    headers = [issues_sheet.cell(row=1, column=index).value for index in range(1, 9)]
    assert headers == [
        "Severidade",
        "Codigo",
        "Mensagem",
        "Como corrigir",
        "Arquivo",
        "Linha",
        "Campo",
        "Contrato",
    ]
    assert issues_sheet["D2"].value == "Preencha o campo obrigatório na linha indicada."
    assert issues_sheet.freeze_panes == "A2"
    assert issues_sheet.auto_filter.ref is not None


def test_export_package_with_attachments_normalizes_clid_layout() -> None:
    dataset = AribaDataset(
        contract_content_documents=[
            {
                "Workspace": "LCW4700001278",
                "File": "Documentos CLID/CLID_4700001278.xlsx",
                "title": "CLID_4700001278.xlsx",
            }
        ]
    )

    attachments_buffer = BytesIO()
    with ZipFile(attachments_buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("Documentos CLID/CLID_4700001278.xlsx", _build_portuguese_clid_workbook())

    package_bytes = build_package_zip_with_attachments(
        dataset=dataset,
        attachments_zip_bytes=attachments_buffer.getvalue(),
        include_report_json=False,
    )

    with ZipFile(BytesIO(package_bytes), "r") as archive:
        clid_bytes = archive.read("Documentos CLID/CLID_4700001278.xlsx")

    workbook = load_workbook(BytesIO(clid_bytes), data_only=False)
    assert workbook.sheetnames == [
        "Contract Header",
        "Contract Item Information",
        "Header Attributes",
        "Item Attributes",
    ]

    contract_header = workbook["Contract Header"]
    contract_header_columns = [contract_header.cell(row=1, column=index).value for index in range(1, 20)]
    assert contract_header_columns[:6] == [
        "Event ID",
        "Title",
        "Description",
        "Requester",
        "Company Name",
        "Supplier Name",
    ]
    assert contract_header["A2"].value == "Doc123"
    assert contract_header["B2"].value == "Contrato teste"

    item_information = workbook["Contract Item Information"]
    item_information_columns = [item_information.cell(row=1, column=index).value for index in range(1, 28)]
    assert item_information_columns[-3:] == [
        "External System Line Number",
        "Source Event ID",
        "Line Item Number",
    ]
    assert item_information["B2"].value == "10"
    assert item_information["Z2"].value == "Doc123"
    assert item_information["AA2"].value == "10"

    item_attributes = workbook["Item Attributes"]
    item_attribute_columns = [item_attributes.cell(row=1, column=index).value for index in range(1, 10)]
    assert item_attribute_columns == [
        "Item Number",
        "Attribute Name",
        "Attribute Value",
        "Display Text",
        "Type",
        "Description",
        "Is Term added from Item Master",
        "Formula",
        "Item Status",
    ]
    assert item_attributes["A2"].value == "10"
    assert item_attributes["B2"].value == "Plant"
